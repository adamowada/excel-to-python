# Imports and Setup
import pandas as pd
from openpyxl import load_workbook
import click
import formulas
import logging
from pathlib import Path  # Use Path for cleaner path handling


# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# CLI Interface with Click
@click.command()
@click.argument('excel_filepath', type=click.Path(exists=True))
def convert_excel_to_python(excel_filepath):
    try:
        # Use pathlib for cleaner path handling
        dataframes = create_dataframes_from_excel(Path(excel_filepath))
        save_output_files(Path(excel_filepath), dataframes)
    except Exception as e:
        logger.error(f"Failed to convert Excel file: {e}")
        raise click.ClickException(f"An error occurred during conversion: {e}")


# Excel Parsing Logic
def parse_excel_file(excel_filepath):
    try:
        # Convert to Path and use `resolve` for full path
        excel_filepath = Path(excel_filepath).resolve()
        workbook = load_workbook(excel_filepath, data_only=False)
        sheets_data = {}
        logger.info(f"Processing Excel file: {excel_filepath}")

        for sheet_name in workbook.sheetnames:
            logger.info(f"Processing sheet: {sheet_name}")
            sheet = workbook[sheet_name]
            sheet_data = extract_formulas_from_sheet(sheet)
            if sheet_data:
                sheets_data[sheet_name] = sheet_data

        return sheets_data

    except FileNotFoundError:
        logger.error(f"Excel file not found: {excel_filepath}")
        raise click.ClickException(f"File not found: {excel_filepath}")
    except Exception as e:
        logger.error(f"Error while parsing Excel file {excel_filepath}: {e}")
        raise click.ClickException(f"Failed to parse Excel file: {e}")


def extract_formulas_from_sheet(sheet):
    formulas_data = {}
    try:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Check if the cell contains a formula
                    formula = cell.value
                    referenced_cells = get_referenced_cells(formula)
                    if referenced_cells is not None:
                        formulas_data[cell.coordinate] = {
                            'formula': formula,
                            'referenced_cells': referenced_cells
                        }
        return formulas_data if formulas_data else None
    except Exception as e:
        logger.error(
            f"Error while extracting formulas from sheet {sheet.title}: {e}")
        raise


def get_referenced_cells(formula):
    try:
        parser = formulas.Parser()
        parsed_formula = parser.ast(formula).parsed
        referenced_cells = []

        for token in parsed_formula:
            if token.ptg == 'operand' and token.subtype == 'range':
                referenced_cells.append(token.value)

        return referenced_cells

    except formulas.exc.FormulaError as e:
        logger.warning(f"Formula parsing error for '{formula}': {e}")
        return None  # Return None if formula parsing fails
    except Exception as e:
        logger.error(f"Unexpected error parsing formula '{formula}': {e}")
        return None  # Catch all unexpected exceptions


# Dataframe Handling
def create_dataframes_from_excel(excel_filepath):
    try:
        sheets_data = parse_excel_file(excel_filepath)
        if not sheets_data:
            raise click.ClickException(
                "No data was extracted from the Excel file.")

        dataframes = {}

        for sheet_name, sheet_data in sheets_data.items():
            sheet_dataframes = organize_data_into_dataframes(sheet_data)
            if sheet_dataframes:  # Ensure non-empty dataframes
                dataframes[sheet_name] = sheet_dataframes

        return dataframes

    except click.ClickException as e:
        raise e
    except Exception as e:
        logger.error(f"Error creating dataframes from Excel file: {e}")
        raise click.ClickException(f"Failed to create dataframes: {e}")


def organize_data_into_dataframes(sheet_data):
    # Create a dataframe directly instead of using an intermediate list
    df_data = {
        'Cell': [],
        'Formula': [],
        'Referenced Cells': []
    }

    for cell, data in sheet_data.items():
        df_data['Cell'].append(cell)
        df_data['Formula'].append(data['formula'])
        df_data['Referenced Cells'].append(', '.join(data['referenced_cells']) if data['referenced_cells'] else 'None')

    return [pd.DataFrame(df_data)]  # Return a list with a single dataframe


def save_dataframes(dataframes, output_dir):
    try:
        # Use pathlib for a cleaner check
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        for sheet_name, sheet_dataframes in dataframes.items():
            for i, dataframe in enumerate(sheet_dataframes):
                if dataframe.empty:
                    logger.info(f"Skipping empty dataframe for {sheet_name}_df{i+1}")
                    continue

                file_path = output_dir / f"{sheet_name}_df{i+1}.csv"
                dataframe.to_csv(file_path, index=False)
                logger.info(f"Saved dataframe to {file_path}")

    except (IOError, OSError) as e:
        logger.error(f"Error saving dataframes to {output_dir}: {e}")
        raise


# Python CLI Code Generation
def generate_python_cli_app(excel_filepath, output_dir):
    try:
        cli_script_path = Path(output_dir) / "main.py"
        dataframes_dir = Path(output_dir) / "dataframes"

        # Move this check earlier to reduce nesting
        if not dataframes_dir.exists() or not list(dataframes_dir.glob("*.csv")):
            raise click.ClickException(f"No dataframes found at {dataframes_dir}")

        with cli_script_path.open('w') as cli_script:
            cli_script.write("# Auto-generated Python CLI based on Excel file\n")
            cli_script.write("import click\n")
            cli_script.write("import pandas as pd\n")
            cli_script.write("from pathlib import Path\n\n")

            cli_script.write("@click.group()\n")
            cli_script.write("def cli():\n")
            cli_script.write("    pass\n\n")

            for sheet_file in dataframes_dir.glob("*.csv"):
                func_name = sheet_file.stem
                logger.info(f"Generating CLI command for {func_name}")
                cli_script.write(f"@cli.command()\n")
                cli_script.write(f"def {func_name}():\n")
                cli_script.write(f"    df_path = Path(__file__).parent / 'dataframes' / '{sheet_file.name}'\n")
                cli_script.write(f"    if not df_path.exists():\n")
                cli_script.write(f"        click.echo('Dataframe file {sheet_file.name} does not exist.')\n")
                cli_script.write(f"        return\n")
                cli_script.write(f"    df = pd.read_csv(df_path)\n")
                cli_script.write(f"    if df.empty:\n")
                cli_script.write(f"        click.echo('No data to display for {func_name}.')\n")
                cli_script.write(f"    else:\n")
                cli_script.write(f"        click.echo(df.to_string())\n\n")

            cli_script.write("if __name__ == '__main__':\n")
            cli_script.write("    cli()\n")

        logger.info(f"Generated Python CLI at {cli_script_path}")

    except click.ClickException as e:
        raise e
    except Exception as e:
        logger.error(f"Error generating Python CLI: {e}")
        raise click.ClickException(f"Failed to generate CLI app: {e}")


def translate_excel_formulas_to_python(formula):
    try:
        parser = formulas.Parser()
        parser.parse(formula)
        parsed_formula = parser.ast
        # Convert parsed AST to Python code
        # Note: This may require manual adjustments for complex Excel formulas
        python_code = parsed_formula.to_python()

        return python_code

    except formulas.exc.FormulaError as e:
        logger.warning(f"Error translating Excel formula '{formula}': {e}")
        return None  # Return None if translation fails
    except Exception as e:
        logger.error(f"Unexpected error translating Excel formula '{formula}': {e}")
        return None  # Catch all unexpected exceptions


def create_menu_for_multiple_functionalities(functionalities):
    @click.group()
    def cli():
        pass

    for func_name, func_action in functionalities.items():
        cli.command(name=func_name)(func_action)

    return cli


# Output Handling
def save_output_files(excel_filepath, dataframes):
    try:
        output_dir = Path("outputs") / Path(excel_filepath).stem

        if not output_dir.exists():
            output_dir.mkdir(parents=True, exist_ok=True)

        if not dataframes:
            raise click.ClickException("No dataframes to save.")

        save_dataframes(dataframes, output_dir / "dataframes")
        generate_python_cli_app(excel_filepath, output_dir)

    except click.ClickException as e:
        raise e
    except (IOError, OSError) as e:
        logger.error(f"Error saving output files: {e}")
        raise click.ClickException(f"Failed to save output files: {e}")


if __name__ == "__main__":
    convert_excel_to_python()
